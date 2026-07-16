"""
Carga automatica dos Pedidos em Transito (Norte) para o Supabase.

Le a planilha diaria (aba "Base de pedidos", cabecalho na linha 5), filtra os
7 vendedores do Norte, converte datas/numeros e faz upsert por `chave`.
Registra o nome do arquivo em `carga_info` (aparece no aviso do Dashboard).

Uso:
    python carga_transito.py              # carga real
    python carga_transito.py --dry-run    # so le e mostra o resumo (nao grava)
    python carga_transito.py --arquivo "C:\\caminho\\Em transito.xlsx"

Config: automacao/config.env (NAO versionado). Veja config.exemplo.env.
Requer: openpyxl  (pip install -r requirements.txt)
"""
import argparse
import glob
import json
import os
import re
import sys
import unicodedata
import urllib.error
import urllib.request
from datetime import date, datetime, timedelta

from openpyxl import load_workbook

AQUI = os.path.dirname(os.path.abspath(__file__))
SHEET = "Base de pedidos"
HEADER_ROW = 5  # 1-indexed: linhas 1-4 sao titulo/metadados

VENDEDORES_NORTE = {
    "FURTADO E GEMAQUE LTDA (FREDERICSON)",
    "NAILSON F COSTA",
    "OREN REPRESENTACOES (ROSIMARA)",
    "DANIELA NASCIMENTO DA SILVA",
    "FURTADO E GEMAQUE LTDA (ANA GEMAQUE)",
    "ORTIZ E OLIVEIRA REP E COM (SCARLETTY)",
    "ES ANDRADE REPRESENTACOES (EDUARDO)",
}

COLUNAS = {
    "Chave": "chave",
    "CNPJ Cliente": "cnpj_cliente",
    "Nome Cliente": "nome_cliente",
    "Vendedor": "vendedor",
    "Cidade": "cidade",
    "Estado": "estado",
    "Região Brasil": "regiao_brasil",
    "Região": "regiao",
    "Filial": "filial",
    "Número Pedido (Protheus)": "numero_pedido",
    "Pedido SalesForce (Oportunidade)": "pedido_salesforce",
    "Ped. Cliente": "pedido_cliente",
    "NF": "nf",
    "Valor Faturado": "valor_faturado",
    "Peso": "peso",
    "Desc. Tipo Saída": "desc_tipo_saida",
    "Operação": "operacao",
    "Transportador": "transportador",
    "Data Faturamento": "data_faturamento",
    "Data programada expedição": "data_prog_expedicao",
    "Data expedição": "data_expedicao",
    "Lead time (dias úteis)": "lead_time",
    "Previsão de entrega": "previsao_entrega",
    "Data Real de Chegada": "data_real_chegada",
    "Data Real Entrega": "data_real_entrega",
    "Recebimento": "recebimento",
    "Data Agenda": "data_agenda",
    "Status trânsito": "status_transito",
    "OBS. TRÂNSITO": "obs_transito",
}

CAMPOS_DATA = {
    "data_faturamento", "data_prog_expedicao", "data_expedicao",
    "previsao_entrega", "data_real_chegada", "data_real_entrega", "data_agenda",
}
CAMPOS_NUM = {"valor_faturado", "peso"}
CAMPOS_INT = {"lead_time"}

LOTE = 500


# ---------------------------------------------------------------- utilidades
def normalizar(h):
    s = unicodedata.normalize("NFD", str(h if h is not None else ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", s).strip().lower()


MAPA = {normalizar(k): v for k, v in COLUNAS.items()}


def parse_data(v):
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        # serial do Excel (base 1899-12-30)
        return (datetime(1899, 12, 30) + timedelta(days=float(v))).date().isoformat()
    s = str(v).strip()
    if not s:
        return None
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", s)
    if m:
        return f"{m.group(3)}-{int(m.group(2)):02d}-{int(m.group(1)):02d}"
    return None


def parse_num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(".", "").replace(",", ".")
    s = re.sub(r"[^0-9.\-]", "", s)
    try:
        return float(s)
    except ValueError:
        return None


def carregar_config():
    cfg = {}
    caminho = os.path.join(AQUI, "config.env")
    if os.path.exists(caminho):
        with open(caminho, encoding="utf-8") as fh:
            for linha in fh:
                linha = linha.strip()
                if not linha or linha.startswith("#") or "=" not in linha:
                    continue
                k, v = linha.split("=", 1)
                cfg[k.strip()] = v.strip()
    for k in ("ARQUIVO_XLSX", "PASTA_XLSX", "PADRAO", "SUPABASE_URL",
              "SUPABASE_ANON_KEY", "ADMIN_EMAIL", "ADMIN_SENHA"):
        if os.environ.get(k):
            cfg[k] = os.environ[k]
    return cfg


def achar_arquivo(cfg, arg_arquivo=None):
    if arg_arquivo:
        alvo = arg_arquivo
    elif cfg.get("ARQUIVO_XLSX"):
        alvo = cfg["ARQUIVO_XLSX"]
    else:
        pasta = cfg.get("PASTA_XLSX")
        if not pasta:
            sys.exit("ERRO: defina ARQUIVO_XLSX ou PASTA_XLSX no automacao/config.env")
        padrao = cfg.get("PADRAO", "*.xlsx")
        cands = [p for p in glob.glob(os.path.join(pasta, padrao))
                 if not os.path.basename(p).startswith("~$")]
        if not cands:
            sys.exit(f"ERRO: nenhum arquivo em {pasta} com padrao {padrao}")
        alvo = max(cands, key=os.path.getmtime)
    if not os.path.exists(alvo):
        sys.exit(f"ERRO: arquivo nao encontrado: {alvo}")
    return alvo


# ------------------------------------------------------------------- parsing
def ler_planilha(caminho):
    wb = load_workbook(caminho, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        sys.exit(f'ERRO: aba "{SHEET}" nao encontrada. Abas: {wb.sheetnames}')
    ws = wb[SHEET]
    linhas = ws.iter_rows(min_row=HEADER_ROW, values_only=True)
    try:
        cabecalho = next(linhas)
    except StopIteration:
        sys.exit("ERRO: planilha vazia na linha de cabecalho")

    campos = [MAPA.get(normalizar(h)) for h in cabecalho]
    if "chave" not in campos or "vendedor" not in campos:
        sys.exit(f"ERRO: cabecalho nao reconhecido na linha {HEADER_ROW}. Lido: {cabecalho}")

    pedidos, ignorados = [], 0
    for linha in linhas:
        ped = {}
        for campo, valor in zip(campos, linha):
            if not campo:
                continue
            if campo in CAMPOS_DATA:
                ped[campo] = parse_data(valor)
            elif campo in CAMPOS_NUM:
                ped[campo] = parse_num(valor)
            elif campo in CAMPOS_INT:
                n = parse_num(valor)
                ped[campo] = None if n is None else int(round(n))
            else:
                s = None if valor is None else str(valor).strip()
                ped[campo] = s if s else None
        if not ped.get("chave"):
            continue
        if ped.get("vendedor") not in VENDEDORES_NORTE:
            ignorados += 1
            continue
        pedidos.append(ped)
    wb.close()
    return pedidos, ignorados


# ------------------------------------------------------------------ supabase
def _req(url, headers, payload=None, metodo="GET"):
    data = json.dumps(payload).encode("utf-8") if payload is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=metodo)
    try:
        with urllib.request.urlopen(req, timeout=60) as resp:
            corpo = resp.read().decode("utf-8")
            return resp.status, corpo
    except urllib.error.HTTPError as e:
        sys.exit(f"ERRO HTTP {e.code} em {url}\n{e.read().decode('utf-8', 'ignore')}")


def login(cfg):
    url = f"{cfg['SUPABASE_URL']}/auth/v1/token?grant_type=password"
    h = {"apikey": cfg["SUPABASE_ANON_KEY"], "Content-Type": "application/json"}
    _, corpo = _req(url, h, {"email": cfg["ADMIN_EMAIL"], "password": cfg["ADMIN_SENHA"]}, "POST")
    dados = json.loads(corpo)
    if not dados.get("access_token"):
        sys.exit(f"ERRO: login falhou: {corpo}")
    return dados["access_token"], dados["user"]["id"]


def chaves_existentes(cfg, token):
    url = f"{cfg['SUPABASE_URL']}/rest/v1/pedidos?select=chave&limit=100000"
    h = {"apikey": cfg["SUPABASE_ANON_KEY"], "Authorization": f"Bearer {token}"}
    _, corpo = _req(url, h)
    return {r["chave"] for r in json.loads(corpo)}


def upsert(cfg, token, pedidos):
    url = f"{cfg['SUPABASE_URL']}/rest/v1/pedidos?on_conflict=chave"
    h = {
        "apikey": cfg["SUPABASE_ANON_KEY"], "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
        "Prefer": "resolution=merge-duplicates,return=minimal",
    }
    carimbo = datetime.now().astimezone().isoformat()
    for i in range(0, len(pedidos), LOTE):
        lote = [dict(p, updated_at=carimbo) for p in pedidos[i:i + LOTE]]
        _req(url, h, lote, "POST")


def registrar_carga(cfg, token, user_id, arquivo, total, inseridos, atualizados, ignorados):
    url = f"{cfg['SUPABASE_URL']}/rest/v1/carga_info"
    h = {
        "apikey": cfg["SUPABASE_ANON_KEY"], "Authorization": f"Bearer {token}",
        "Content-Type": "application/json", "Prefer": "return=minimal",
    }
    _req(url, h, {
        "arquivo": os.path.basename(arquivo), "total": total, "inseridos": inseridos,
        "atualizados": atualizados, "ignorados": ignorados, "user_id": user_id,
    }, "POST")


# ---------------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser(description="Carga dos Pedidos em Transito (Norte)")
    ap.add_argument("--dry-run", action="store_true", help="so le e mostra o resumo")
    ap.add_argument("--arquivo", help="caminho do .xlsx (sobrescreve o config)")
    args = ap.parse_args()

    cfg = carregar_config()
    arquivo = achar_arquivo(cfg, args.arquivo)
    print(f"Arquivo: {arquivo}")
    print(f"Modificado em: {datetime.fromtimestamp(os.path.getmtime(arquivo)):%d/%m/%Y %H:%M}")

    pedidos, ignorados = ler_planilha(arquivo)
    print(f"Lidos: {len(pedidos)} pedidos do Norte | {ignorados} ignorados (outras regioes)")
    if pedidos:
        vend = {}
        for p in pedidos:
            vend[p["vendedor"]] = vend.get(p["vendedor"], 0) + 1
        for v, c in sorted(vend.items(), key=lambda x: -x[1]):
            print(f"   {c:>5}  {v}")

    if args.dry_run:
        print("\n[dry-run] Nada foi gravado.")
        return
    if not pedidos:
        sys.exit("Nada para gravar.")

    for k in ("SUPABASE_URL", "SUPABASE_ANON_KEY", "ADMIN_EMAIL", "ADMIN_SENHA"):
        if not cfg.get(k):
            sys.exit(f"ERRO: falta {k} no automacao/config.env")

    token, user_id = login(cfg)
    antigas = chaves_existentes(cfg, token)
    novas = [p["chave"] for p in pedidos]
    atualizados = sum(1 for c in novas if c in antigas)
    inseridos = len(novas) - atualizados

    upsert(cfg, token, pedidos)
    registrar_carga(cfg, token, user_id, arquivo, len(novas), inseridos, atualizados, ignorados)

    print(f"\nOK! Inseridos: {inseridos} | Atualizados: {atualizados} | Total: {len(novas)}")
    print(f"Registrado no app como: {os.path.basename(arquivo)}")


if __name__ == "__main__":
    main()
